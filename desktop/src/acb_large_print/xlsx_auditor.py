"""Audit Excel workbooks for accessibility compliance."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

try:
    from openpyxl.chart import ChartBase
except ImportError:
    from openpyxl.chart._chart import ChartBase
from openpyxl.worksheet.worksheet import Worksheet

from . import constants as C
from .auditor import AuditResult

# Generic default sheet names across locales
_GENERIC_SHEET_NAMES = frozenset(
    {
        "Sheet",
        "Sheet1",
        "Sheet2",
        "Sheet3",
        "Sheet4",
        "Sheet5",
        "Tabelle1",
        "Tabelle2",
        "Feuil1",
        "Feuil2",
        "Hoja1",
        "Hoja2",
        "Foglio1",
        "Foglio2",
    }
)

# Non-descriptive link text patterns (shared with Word auditor)
_BAD_LINK_RE = re.compile(
    r"^(click here|here|link|more|read more|learn more|download|info)$",
    re.IGNORECASE,
)
_URL_RE = re.compile(r"^https?://", re.IGNORECASE)


def audit_workbook(file_path: str | Path) -> AuditResult:
    """Audit an Excel workbook for accessibility issues.

    Returns an AuditResult populated with findings scoped to XLSX rules.
    """
    file_path = Path(file_path)
    result = AuditResult(file_path=str(file_path))

    wb = load_workbook(str(file_path), data_only=True)

    # -- Document properties --
    if not (wb.properties.title and wb.properties.title.strip()):
        result.add("XLSX-TITLE", "Workbook title is not set in document properties.")

    if not (wb.properties.creator and wb.properties.creator.strip()):
        result.add(
            "ACB-DOC-AUTHOR", "Workbook author is not set in document properties."
        )

    # -- Per-sheet checks --
    for ws in wb.worksheets:
        sheet_loc = f"Sheet: {ws.title}"

        # Sheet name checks
        if ws.title in _GENERIC_SHEET_NAMES:
            result.add(
                "XLSX-SHEET-NAME",
                f"Worksheet tab '{ws.title}' has a default/generic name.",
                sheet_loc,
            )

        if len(ws.title) > 31:
            result.add(
                "XLSX-SHEET-NAME-LENGTH",
                f"Worksheet tab name '{ws.title}' is {len(ws.title)} characters (max 31).",
                sheet_loc,
            )

        # Merged cells
        if ws.merged_cells.ranges:
            for merged in ws.merged_cells.ranges:
                result.add(
                    "XLSX-MERGED-CELLS",
                    f"Merged range {merged} disrupts screen reader navigation.",
                    sheet_loc,
                )

        # Hidden rows / columns
        _check_hidden_content(ws, result, sheet_loc)

        # Header row frozen
        _check_header_frozen(ws, result, sheet_loc)

        # Charts
        _check_charts(ws, result, sheet_loc)

        # Hyperlinks
        _check_hyperlinks(ws, result, sheet_loc)

        # Color-only cells
        _check_color_only(ws, result, sheet_loc)

        # Empty table cells in data region
        _check_empty_cells(ws, result, sheet_loc)

        # Blank / generic column headers
        _check_column_headers(ws, result, sheet_loc)

        # Blank rows used for layout
        _check_blank_rows_layout(ws, result, sheet_loc)

    # -- Excel Table objects --
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            _check_table_object(tbl, ws, result)

    # -- Images / shapes alt text --
    for ws in wb.worksheets:
        _check_images(ws, result, f"Sheet: {ws.title}")

    wb.close()

    # Sort findings by severity
    severity_order = {
        C.Severity.CRITICAL: 0,
        C.Severity.HIGH: 1,
        C.Severity.MEDIUM: 2,
        C.Severity.LOW: 3,
    }
    result.findings.sort(key=lambda f: severity_order.get(f.severity, 4))

    return result


# ---------------------------------------------------------------------------
# Individual check helpers
# ---------------------------------------------------------------------------


def _check_hidden_content(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Flag hidden rows and columns."""
    hidden_cols = [
        c for c in (ws.column_dimensions or {}).values() if getattr(c, "hidden", False)
    ]
    if hidden_cols:
        result.add(
            "XLSX-HIDDEN-CONTENT",
            f"{len(hidden_cols)} hidden column(s) may be missed by screen readers.",
            loc,
        )

    hidden_rows = [
        r for r in (ws.row_dimensions or {}).values() if getattr(r, "hidden", False)
    ]
    if hidden_rows:
        result.add(
            "XLSX-HIDDEN-CONTENT",
            f"{len(hidden_rows)} hidden row(s) may be missed by screen readers.",
            loc,
        )


def _check_header_frozen(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Check if header row is frozen on sheets with meaningful data."""
    if ws.max_row and ws.max_row > 10:
        freeze = ws.freeze_panes
        if not freeze or freeze == "A1":
            result.add(
                "XLSX-HEADER-FROZEN",
                "Header row is not frozen; column context scrolls out of view.",
                loc,
            )


def _check_charts(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Check charts for alt text."""
    for chart in ws._charts:
        title_text = ""
        if hasattr(chart, "title") and chart.title:
            title_text = str(chart.title)
        if not title_text.strip():
            result.add(
                "PPTX-CHART-ALT-TEXT",
                "A chart has no title or alt text.",
                loc,
            )


def _check_hyperlinks(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Check hyperlink display text."""
    for link in getattr(ws, "_hyperlinks", []):
        display = getattr(link, "display", "") or ""
        target = getattr(link, "target", "") or ""
        if not display.strip():
            continue
        if _BAD_LINK_RE.match(display.strip()):
            result.add(
                "ACB-LINK-TEXT",
                f"Non-descriptive link text '{display.strip()}'.",
                loc,
            )
        elif _URL_RE.match(display.strip()):
            result.add(
                "ACB-LINK-TEXT",
                f"Raw URL used as link text: '{display.strip()[:60]}...'.",
                loc,
            )


def _check_color_only(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Flag cells with background color but no text (color-only meaning)."""
    flagged = 0
    max_check = min(ws.max_row or 0, 200)  # limit for performance
    max_col = min(ws.max_column or 0, 50)
    for row in ws.iter_rows(min_row=1, max_row=max_check, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                continue
            fill = cell.fill
            if (
                fill
                and fill.fgColor
                and fill.fgColor.rgb
                and fill.fgColor.rgb
                not in (
                    "00000000",
                    None,
                    "0",
                )
            ):
                flagged += 1
    if flagged:
        result.add(
            "XLSX-COLOR-ONLY",
            f"{flagged} empty cell(s) with background color may convey meaning through color alone.",
            loc,
        )


def _check_empty_cells(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Flag empty cells within the apparent data region."""
    max_check = min(ws.max_row or 0, 200)
    max_col = min(ws.max_column or 0, 50)
    empty_count = 0
    for row in ws.iter_rows(min_row=2, max_row=max_check, max_col=max_col):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                empty_count += 1
    if empty_count > 5:
        result.add(
            "ACB-EMPTY-TABLE-CELL",
            f"{empty_count} empty cells in the data region. Consider using a dash or N/A.",
            loc,
        )


def _check_column_headers(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Check first row for blank or generic column headers."""
    if not ws.max_column or ws.max_row == 0:
        return
    max_col = min(ws.max_column, 50)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        val = str(cell.value).strip() if cell.value is not None else ""
        if not val:
            result.add(
                "XLSX-BLANK-COLUMN-HEADER",
                f"Column {col_idx} header is blank.",
                loc,
            )
        elif re.match(r"^Column\s*\d+$", val, re.IGNORECASE):
            result.add(
                "XLSX-BLANK-COLUMN-HEADER",
                f"Column {col_idx} header '{val}' is a generic label.",
                loc,
            )


def _check_table_object(tbl, ws: Worksheet, result: AuditResult) -> None:
    """Check Excel Table objects for header row and default names."""
    loc = f"Sheet: {ws.title}, Table: {tbl.name}"

    # Default table name
    if re.match(r"^Table\d+$", tbl.name or ""):
        result.add(
            "XLSX-DEFAULT-TABLE-NAME",
            f"Excel Table '{tbl.name}' has a default name. Rename to describe its content.",
            loc,
        )

    # Header row enabled check (headerRowCount is None or 0 if disabled)
    if hasattr(tbl, "headerRowCount") and tbl.headerRowCount == 0:
        result.add(
            "XLSX-TABLE-HEADERS",
            f"Excel Table '{tbl.name}' has the header row turned off.",
            loc,
        )


def _check_blank_rows_layout(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Flag runs of 3+ consecutive fully-blank rows used for visual spacing."""
    max_row = ws.max_row or 0
    max_col = min(ws.max_column or 1, 50)
    if max_row < 3:
        return

    consecutive_blank = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            consecutive_blank += 1
            if consecutive_blank == 3:
                result.add(
                    "XLSX-BLANK-ROWS-LAYOUT",
                    "3 or more consecutive blank rows detected. "
                    "Screen readers announce blank cells, creating a confusing reading experience. "
                    "Remove blank rows used for visual spacing.",
                    loc,
                )
                # One finding per sheet is enough
                return
        else:
            consecutive_blank = 0


def _check_images(ws: Worksheet, result: AuditResult, loc: str) -> None:
    """Check images in worksheet for alt text."""
    for img in ws._images:
        desc = ""
        if hasattr(img, "desc"):
            desc = img.desc or ""
        elif hasattr(img, "_element"):
            desc = img._element.get("descr", "") or ""
        if not desc.strip():
            result.add(
                "ACB-MISSING-ALT-TEXT",
                "An image has no alternative text.",
                loc,
            )
