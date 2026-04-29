"""Excel accessibility fixer — named-range header annotation.

Implements an openpyxl-based equivalent of Jamal Mazrui's xlHeaders VBScript
tool (https://github.com/jamalmazrui/xlHeaders).  Detects tabular regions in
each worksheet and creates DefinedName entries that assistive technology (e.g.
JAWS, NVDA + Excel) can use to announce column and row header context.

Named range conventions (matching xlHeaders):
  ColumnTitle01 … ColumnTitleNN   — header cell for column N
  RowTitle01 … RowTitleNN         — first-column header for row N
  Title01                         — sheet/table-level title cell
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


def add_excel_header_named_ranges(workbook_path: str | Path) -> int:
    """Add xlHeaders-style DefinedName entries to an Excel workbook.

    Opens *workbook_path*, detects tabular regions in each worksheet, and
    creates named ranges that identify column-header cells, row-header cells,
    and sheet-level title cells.  Saves the workbook in place.

    Returns the total count of named ranges added across all sheets.
    """
    workbook_path = Path(workbook_path)
    wb = load_workbook(str(workbook_path))
    total_added = 0

    for ws in wb.worksheets:
        total_added += _annotate_sheet(wb, ws)

    wb.save(str(workbook_path))
    wb.close()
    return total_added


def _annotate_sheet(wb, ws) -> int:
    """Detect tabular region in *ws* and add DefinedNames to *wb*.

    Returns count of names added.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Need at least a header row + one data row, and at least two columns
    if max_row < 2 or max_col < 2:
        return 0

    # Read the first row to check for column headers
    header_row = [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, max_col + 1)
    ]

    # Require that the majority of header cells are non-empty and unique
    non_empty = [h for h in header_row if h]
    if len(non_empty) < max(2, max_col // 2):
        return 0
    if len(set(non_empty)) < len(non_empty):
        # Duplicate headers — skip (can't safely pick which one to use)
        return 0

    sheet_name = ws.title
    # openpyxl DefinedName uses the quoted-sheet-name form
    quoted = _quote_sheet_name(sheet_name)
    added = 0

    # Title01 — use the first non-empty header cell (or cell A1) as the title
    title_col = next((i + 1 for i, h in enumerate(header_row) if h), 1)
    title_addr = f"{quoted}!{_cell_addr(1, title_col)}"
    _add_name(wb, f"Title01", title_addr)
    added += 1

    # ColumnTitleNN — one name per header column (1-based, zero-padded to 2 digits)
    for col_idx, header in enumerate(header_row, 1):
        if not header:
            continue
        name = f"ColumnTitle{col_idx:02d}"
        addr = f"{quoted}!{_cell_addr(1, col_idx)}"
        _add_name(wb, name, addr)
        added += 1

    # RowTitleNN — if the first column values look like row labels (non-numeric, unique)
    first_col_values = [
        str(ws.cell(row=r, column=1).value or "").strip()
        for r in range(2, max_row + 1)
    ]
    non_empty_row_labels = [v for v in first_col_values if v]
    all_non_numeric = all(not _looks_numeric(v) for v in non_empty_row_labels)

    if (
        non_empty_row_labels
        and all_non_numeric
        and len(set(non_empty_row_labels)) == len(non_empty_row_labels)
    ):
        for row_offset, label in enumerate(first_col_values, 2):
            if not label:
                continue
            name = f"RowTitle{(row_offset - 1):02d}"
            addr = f"{quoted}!{_cell_addr(row_offset, 1)}"
            _add_name(wb, name, addr)
            added += 1

    return added


def _add_name(wb, name: str, addr: str) -> None:
    """Add a DefinedName to the workbook, silently replacing any existing entry."""
    # openpyxl uses a dict-like store; delete first to avoid collision
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except (KeyError, TypeError):
        pass
    dn = DefinedName(name=name, attr_text=addr)
    wb.defined_names[name] = dn


def _cell_addr(row: int, col: int) -> str:
    """Convert 1-based row/col to an absolute cell address like $A$1."""
    col_letters = _col_letter(col)
    return f"${col_letters}${row}"


def _col_letter(col: int) -> str:
    """Convert 1-based column index to Excel column letter(s)."""
    result = ""
    while col:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _quote_sheet_name(name: str) -> str:
    """Wrap sheet name in single quotes if it contains spaces or special chars."""
    if " " in name or any(c in name for c in "![]*/\\?:"):
        return f"'{name}'"
    return name


def _looks_numeric(value: str) -> bool:
    """Return True if the value looks like a plain number."""
    try:
        float(value.replace(",", "").replace("$", "").replace("%", ""))
        return True
    except ValueError:
        return False
